import openpyxl
wk= openpyxl.Workbook()
#print(wk.active.title)
ws=wk.active
ws.title= "Ram"
print(ws.title)

ws['A1']="RamadhasManimaran"

#SecondSheet creation
wk.create_sheet(title="Mani")
ws1=wk['Mani']
ws1['A3']="8098984980"

wk.save("C:/Users/ramadhma/Selenium/Ram1.xlsx")
