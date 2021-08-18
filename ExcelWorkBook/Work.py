import openpyxl

# Creating object of it
wk= openpyxl.load_workbook("C:/Users/ramadhma/Selenium/WorkBook.xlsx")

print(wk.sheetnames)
# Fetch th active sheet in the workbook
print(wk.active.title)

# Fetch the Sheet to work on
ws=wk['Ram']
print(ws.title)

#print(ws['A1'].value)
#print(ws['A3'].value)

#wc=ws.cell(1,1)
#print(wc.value)

# By passing argument in a Cell
wc=ws.cell(column=1, row=2)
print(wc.value)
print(wc.row)
print(wc.column)