import openpyxl

def dataGenerator():
    wk= openpyxl.load_workbook("C://Users//ramadhma//Selenium//WorkBook.xlsx")
    ws=wk['Ram']
    r= ws.max_row
    li=[]
    li1=[]
    for i in range (1,r+1):
        li1=[]
        un=ws.cell(i,1)
        up=ws.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,up.value)
        li.insert(i,li1)
    print(li)
    return li


