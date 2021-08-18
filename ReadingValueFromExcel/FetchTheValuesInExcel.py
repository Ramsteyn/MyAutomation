import openpyxl
wk= openpyxl.load_workbook("C:/Users/ramadhma/Selenium/WorkBook.xlsx")
sheetName=wk.active.title
print(sheetName)
ws=wk[sheetName]
rows= ws.max_row
column= ws.max_column

print(rows)
print(column)

#for i in range(1,rows+1):
#    for j in range(1,column+1):
#        wc= ws.cell(i,j)
#        print(wc.value, end="")

# Another method
for r in ws['A1': 'C3']:
    for c in r:
        print(c.value)